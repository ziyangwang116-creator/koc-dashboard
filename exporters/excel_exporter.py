from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any
from unicodedata import east_asian_width

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.transformer import OUTPUT_COLUMNS
from models.koc import KOC_EXPORT_COLUMNS


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NUMBER_FORMAT = "#,##0"
DATE_FORMAT = "yyyy-mm-dd"

PREFERRED_WIDTHS = {
    "koc_name": (14, 24),
    "platform": (12, 18),
    "publish_date": (14, 14),
    "title": (30, 60),
    "url": (48, 70),
    "views": (12, 18),
    "remark": (14, 30),
    "likes": (12, 16),
    "comment": (12, 16),
    "reposted": (12, 16),
    "source_file": (24, 50),
    "error_message": (28, 70),
    "issue_type": (22, 28),
    "detail": (28, 70),
    "user_id": (16, 28),
    "creator_category": (18, 22),
    "contract_type": (18, 22),
    "contract_start_date": (16, 18),
    "contract_end_date": (16, 18),
    "homepage_url": (48, 70),
    "follower_count": (16, 20),
    "follower_raw_display_value": (20, 28),
    "follower_source": (20, 24),
    "follower_source_url": (48, 70),
    "follower_count_is_estimated": (24, 28),
    "follower_count_updated_at": (24, 26),
    "follower_sync_status": (20, 24),
    "settlement_eligible": (20, 24),
    "active": (10, 12),
    "note": (20, 50),
    "created_at": (20, 22),
    "updated_at": (20, 22),
}


def _excel_value(value: Any) -> Any:
    if value is None or pd.isna(value):
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except ValueError:
            pass
    return value


def _write_dataframe(worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
    worksheet.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        worksheet.append([_excel_value(value) for value in row])

    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 24


def _visual_length(value: Any) -> int:
    text = str(value)
    return sum(2 if east_asian_width(character) in {"W", "F"} else 1 for character in text)


def _set_column_widths(worksheet: Worksheet) -> None:
    for column_number, header_cell in enumerate(worksheet[1], start=1):
        header = str(header_cell.value)
        values = [worksheet.cell(row=row, column=column_number).value for row in range(1, worksheet.max_row + 1)]
        content_width = max(
            (_visual_length(value) for value in values if value is not None), default=0
        ) + 2
        minimum, maximum = PREFERRED_WIDTHS.get(header, (12, 36))
        worksheet.column_dimensions[get_column_letter(column_number)].width = min(
            max(content_width, minimum), maximum
        )


def _format_result_sheet(worksheet: Worksheet) -> None:
    header_positions = {cell.value: cell.column for cell in worksheet[1]}
    for column_name in ("views", "likes", "comment"):
        column_number = header_positions[column_name]
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column_number).number_format = NUMBER_FORMAT

    date_column = header_positions["publish_date"]
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=date_column)
        if isinstance(cell.value, (datetime, date)):
            cell.number_format = DATE_FORMAT

    title_column = header_positions["title"]
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=title_column)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if cell.value is not None and _visual_length(cell.value) > 60:
            worksheet.row_dimensions[row].height = 32

    url_column = header_positions["url"]
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=url_column)
        if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"


def _format_koc_master_sheet(worksheet: Worksheet) -> None:
    headers = {cell.value: cell.column for cell in worksheet[1]}
    uid_column = headers["user_id"]
    follower_column = headers["follower_count"]
    homepage_column = headers["homepage_url"]
    source_url_column = headers["follower_source_url"]
    date_columns = [
        headers["contract_start_date"],
        headers["contract_end_date"],
        headers["follower_count_updated_at"],
        headers["created_at"],
        headers["updated_at"],
    ]
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row, column=uid_column).number_format = "@"
        worksheet.cell(row=row, column=follower_column).number_format = NUMBER_FORMAT
        for column in date_columns:
            cell = worksheet.cell(row=row, column=column)
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif isinstance(cell.value, date):
                cell.number_format = DATE_FORMAT
        homepage_cell = worksheet.cell(row=row, column=homepage_column)
        if isinstance(homepage_cell.value, str) and homepage_cell.value.startswith(
            ("http://", "https://")
        ):
            homepage_cell.hyperlink = homepage_cell.value
            homepage_cell.style = "Hyperlink"
        source_url_cell = worksheet.cell(row=row, column=source_url_column)
        if isinstance(source_url_cell.value, str) and source_url_cell.value.startswith(
            ("http://", "https://")
        ):
            source_url_cell.hyperlink = source_url_cell.value
            source_url_cell.style = "Hyperlink"


def export_to_excel(result_data: pd.DataFrame, exceptions: pd.DataFrame) -> bytes:
    if list(result_data.columns) != OUTPUT_COLUMNS:
        raise ValueError("导出数据字段或顺序不符合标准格式。")

    workbook = Workbook()
    result_sheet = workbook.active
    result_sheet.title = "KOC数据"
    _write_dataframe(result_sheet, result_data)
    _set_column_widths(result_sheet)
    _format_result_sheet(result_sheet)

    exception_sheet = workbook.create_sheet("异常信息")
    _write_dataframe(exception_sheet, exceptions)
    _set_column_widths(exception_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_multi_file_excel(
    result_data: pd.DataFrame,
    file_reports: pd.DataFrame,
    exceptions: pd.DataFrame,
) -> bytes:
    """Export the V0.2 batch result using the three required worksheets."""
    if list(result_data.columns) != OUTPUT_COLUMNS:
        raise ValueError("导出数据字段或顺序不符合标准格式。")

    workbook = Workbook()
    result_sheet = workbook.active
    result_sheet.title = "整理结果"
    _write_dataframe(result_sheet, result_data)
    _set_column_widths(result_sheet)
    _format_result_sheet(result_sheet)

    report_sheet = workbook.create_sheet("文件处理报告")
    _write_dataframe(report_sheet, file_reports)
    _set_column_widths(report_sheet)

    exception_sheet = workbook.create_sheet("异常数据")
    exception_data = exceptions.copy()
    if exception_data.empty:
        exception_data = pd.DataFrame(
            [
                {
                    column: "未发现异常" if column == "detail" else None
                    for column in exceptions.columns
                }
            ],
            columns=exceptions.columns,
        )
    _write_dataframe(exception_sheet, exception_data)
    _set_column_widths(exception_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_koc_master(dataframe: pd.DataFrame) -> bytes:
    if list(dataframe.columns) != KOC_EXPORT_COLUMNS:
        raise ValueError("达人库导出字段或顺序不符合新版格式。")
    export_data = dataframe.copy()
    export_data["user_id"] = export_data["user_id"].astype("string")
    export_data["follower_count"] = pd.to_numeric(
        export_data["follower_count"], errors="coerce"
    ).astype("Int64")
    for column in ("contract_start_date", "contract_end_date"):
        export_data[column] = pd.to_datetime(
            export_data[column], errors="coerce"
        ).dt.date
    for column in ("follower_count_updated_at", "created_at", "updated_at"):
        converted = pd.to_datetime(export_data[column], errors="coerce", utc=True)
        export_data[column] = [
            (
                timestamp.to_pydatetime().replace(tzinfo=None)
                if not pd.isna(timestamp)
                else None
            )
            for timestamp in converted
        ]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "达人库"
    _write_dataframe(worksheet, export_data)
    _set_column_widths(worksheet)
    _format_koc_master_sheet(worksheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_download_filename(now: datetime | None = None) -> str:
    timestamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    return f"KOC_整理结果_{timestamp}.xlsx"


def build_multi_file_download_filename(now: datetime | None = None) -> str:
    timestamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    return f"KOC_多文件整理结果_{timestamp}.xlsx"


def build_koc_master_filename(now: datetime | None = None) -> str:
    timestamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    return f"KOC达人库_{timestamp}.xlsx"
