from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from core.file_processor import UploadedExcel
from core.multi_file_processor import MultiFileProcessor
from database.koc_repository import KOCRepository
from exporters.excel_exporter import export_koc_master, export_multi_file_excel


def _rapid_query_file(name: str, user_id: int, url: str, title: str) -> UploadedExcel:
    timestamp = pd.Timestamp("2024-04-01T16:00:00Z").value // 1_000_000
    dataframe = pd.DataFrame(
        {
            "view": [100],
            "subtype": ["short"],
            "title": [title],
            "userId": [user_id],
            "platform": ["YouTube"],
            "url": [url],
            "timestamp": [timestamp],
        }
    )
    output = BytesIO()
    dataframe.to_excel(output, index=False, engine="openpyxl")
    return UploadedExcel(name=name, content=output.getvalue())


def test_v02_complete_local_workflow_with_two_excel_files(tmp_path):
    database_path = tmp_path / "koc.db"
    repository = KOCRepository(database_path)
    created = repository.create(user_id="930001", koc_name="更新前")
    updated = repository.update(
        created.id,
        user_id="930001",
        koc_name="テスト达人",
        creator_category=None,
        contract_type=None,
        homepage_url=None,
        follower_count=None,
        active=True,
        note="端到端验收",
    )
    files = [
        _rapid_query_file("week_1.xlsx", 930001, "https://example.com/1", "第一条"),
        _rapid_query_file("week_2.xlsx", 930001, "https://example.com/2", "第二条"),
    ]

    result = MultiFileProcessor(database_path, "Asia/Shanghai").process(files)
    workbook = load_workbook(
        BytesIO(
            export_multi_file_excel(
                result.data, result.file_reports, result.exceptions
            )
        )
    )

    assert len(result.data) == 2
    assert result.data["koc_name"].tolist() == ["テスト达人", "テスト达人"]
    assert result.data["platform"].tolist() == ["shorts", "shorts"]
    assert workbook.sheetnames == ["整理结果", "文件处理报告", "异常数据"]
    assert workbook["整理结果"]["A2"].value == "テスト达人"

    repository.set_active(updated.id, False)
    after_deactivation = MultiFileProcessor(database_path, "Asia/Shanghai").process(
        [files[0]]
    )
    assert pd.isna(after_deactivation.data.loc[0, "koc_name"])

    backup = load_workbook(BytesIO(export_koc_master(repository.to_dataframe())))
    assert "テスト达人" in [cell.value for cell in backup["达人库"]["B"]]
