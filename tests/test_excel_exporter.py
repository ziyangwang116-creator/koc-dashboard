from datetime import date, datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from core.transformer import OUTPUT_COLUMNS
from exporters.excel_exporter import (
    build_download_filename,
    build_koc_master_filename,
    build_multi_file_download_filename,
    export_multi_file_excel,
    export_to_excel,
)


def test_export_has_expected_sheets_values_and_formatting():
    result_data = pd.DataFrame(
        [["なみかりちゃんねる", "shorts", date(2024, 1, 2), "标题", "https://example.com/video", 123, pd.NA, 5, pd.NA, pd.NA]],
        columns=OUTPUT_COLUMNS,
    )
    exceptions = pd.DataFrame(
        [["未匹配 UID", "999", 2, "待检查"]],
        columns=["异常类型", "userId", "原始行号", "说明"],
    )

    workbook = load_workbook(BytesIO(export_to_excel(result_data, exceptions)))
    sheet = workbook["KOC数据"]

    assert workbook.sheetnames == ["KOC数据", "异常信息"]
    assert [cell.value for cell in sheet[1]] == OUTPUT_COLUMNS
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:J2"
    assert sheet["A1"].font.bold is True
    assert sheet["C2"].number_format == "yyyy-mm-dd"
    assert sheet["D2"].alignment.wrap_text is True
    assert sheet["F2"].number_format == "#,##0"
    assert sheet["G2"].value is None
    assert sheet["I2"].value is None
    assert sheet.column_dimensions["A"].width >= 20
    assert sheet.column_dimensions["E"].width >= 48
    assert workbook["异常信息"]["B2"].value == "999"


def test_download_filename_uses_required_pattern():
    assert build_download_filename(datetime(2025, 2, 3, 4, 5, 6)) == "KOC_整理结果_20250203_040506.xlsx"


def test_v02_export_has_three_sheets_ordered_columns_and_japanese_text():
    result_data = pd.DataFrame(
        [["なみかりちゃんねる", "shorts", date(2024, 1, 2), "日本語タイトル", "https://example.com/a", 123, pd.NA, pd.NA, pd.NA, pd.NA]],
        columns=OUTPUT_COLUMNS,
    )
    reports = pd.DataFrame(
        [["week.xlsx", 1, 1, 0, 0, "成功", ""]],
        columns=["source_file", "original_rows", "processed_rows", "unmatched_uid", "duplicate_url", "status", "error_message"],
    )
    exceptions = pd.DataFrame(
        columns=["issue_type", "source_file", "userId", "koc_name", "publish_date", "title", "url", "detail"]
    )

    workbook = load_workbook(
        BytesIO(export_multi_file_excel(result_data, reports, exceptions))
    )

    assert workbook.sheetnames == ["整理结果", "文件处理报告", "异常数据"]
    result_sheet = workbook["整理结果"]
    assert [cell.value for cell in result_sheet[1]] == OUTPUT_COLUMNS
    assert result_sheet["A2"].value == "なみかりちゃんねる"
    assert result_sheet["D2"].value == "日本語タイトル"
    assert result_sheet.freeze_panes == "A2"
    assert workbook["异常数据"]["H2"].value == "未发现异常"


def test_v02_filenames_use_required_patterns():
    now = datetime(2025, 2, 3, 4, 5, 6)
    assert build_multi_file_download_filename(now) == "KOC_多文件整理结果_20250203_040506.xlsx"
    assert build_koc_master_filename(now) == "KOC达人库_20250203_040506.xlsx"
