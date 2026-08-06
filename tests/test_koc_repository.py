from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.koc_mapper import KOCMapper
from database.db import init_db
from database.koc_repository import DuplicateUserIDError, KOCRepository, KOCRepositoryError
from exporters.excel_exporter import export_koc_master
from models.enums import ContractType, CreatorCategory, FollowerSyncStatus
from models.koc import KOC_EXPORT_COLUMNS


def test_create_koc_success_and_trim_values(tmp_path):
    repository = KOCRepository(tmp_path / "koc.db")

    record = repository.create(
        user_id=" 900001 ",
        koc_name=" テスト达人 ",
        creator_category=CreatorCategory.GRASSROOT,
        contract_type=ContractType.TT,
        homepage_url=" https://www.tiktok.com/@creator ",
        follower_count="12,345",
        note=" 备注 ",
    )

    assert record.user_id == "900001"
    assert record.koc_name == "テスト达人"
    assert record.creator_category is CreatorCategory.GRASSROOT
    assert record.contract_type is ContractType.TT
    assert record.homepage_url == "https://www.tiktok.com/@creator"
    assert record.follower_count == 12345
    assert record.follower_sync_status is FollowerSyncStatus.MANUAL
    assert record.follower_count_updated_at is not None
    assert record.note == "备注"


def test_duplicate_user_id_is_rejected(tmp_path):
    repository = KOCRepository(tmp_path / "koc.db")
    repository.create(user_id="900002", koc_name="A")

    with pytest.raises(DuplicateUserIDError, match="该 UID 已存在"):
        repository.create(user_id="900002.0", koc_name="B")


def test_updated_koc_is_used_by_fresh_mapper_immediately(tmp_path):
    database_path = tmp_path / "koc.db"
    repository = KOCRepository(database_path)
    record = repository.create(user_id="900003", koc_name="修改前")
    repository.update(
        record.id,
        user_id="900004",
        koc_name="修改后",
        creator_category=CreatorCategory.LONG_TERM,
        contract_type=None,
        homepage_url=None,
        follower_count=None,
        active=True,
        note=None,
    )

    names, normalized = KOCMapper.from_database(database_path).map_series(
        pd.Series([900004.0, "900003"])
    )

    assert normalized.tolist() == ["900004", "900003"]
    assert names.iloc[0] == "修改后"
    assert pd.isna(names.iloc[1])


def test_inactive_koc_is_retained_but_not_matched_by_default(tmp_path):
    database_path = tmp_path / "koc.db"
    repository = KOCRepository(database_path)
    record = repository.create(user_id="900005", koc_name="停用达人")
    repository.set_active(record.id, False)

    assert repository.get(record.id).active is False
    names, _ = KOCMapper.from_database(database_path).map_series(pd.Series(["900005"]))
    assert pd.isna(names.iloc[0])


def test_initialization_is_idempotent_and_does_not_overwrite_user_edits(tmp_path):
    database_path = tmp_path / "koc.db"
    repository = KOCRepository(database_path)
    existing = next(row for row in repository.list() if row.user_id == "107258")
    repository.update(
        existing.id,
        user_id=existing.user_id,
        koc_name="用户修改名称",
        creator_category=CreatorCategory.COMMENTARY,
        contract_type=None,
        homepage_url=None,
        follower_count=None,
        active=True,
        note="保留",
    )

    init_db(database_path)
    reloaded = KOCRepository(database_path).get(existing.id)

    assert reloaded.koc_name == "用户修改名称"
    assert reloaded.creator_category is CreatorCategory.COMMENTARY
    assert reloaded.note == "保留"


def test_seed_is_not_reinserted_after_user_changes_seed_uid(tmp_path):
    database_path = tmp_path / "koc.db"
    repository = KOCRepository(database_path)
    existing = next(row for row in repository.list() if row.user_id == "107258")
    repository.update(
        existing.id,
        user_id="107258-new",
        koc_name=existing.koc_name,
        creator_category=None,
        contract_type=None,
        homepage_url=None,
        follower_count=None,
        active=True,
        note=None,
    )

    init_db(database_path)
    all_uids = {row.user_id for row in KOCRepository(database_path).list()}

    assert "107258-new" in all_uids
    assert "107258" not in all_uids


def test_negative_follower_count_is_rejected(tmp_path):
    repository = KOCRepository(tmp_path / "koc.db")
    with pytest.raises(KOCRepositoryError, match="粉丝数"):
        repository.create(user_id="920100", koc_name="负数", follower_count=-1)


def test_koc_master_export_preserves_text_and_new_column_order(tmp_path):
    repository = KOCRepository(tmp_path / "koc.db")
    repository.create(
        user_id="920000",
        koc_name="なみかりちゃんねる😀",
        homepage_url="https://www.youtube.com/@creator",
        follower_count=123456,
    )

    workbook = load_workbook(BytesIO(export_koc_master(repository.to_dataframe())))
    sheet = workbook["达人库"]

    assert [cell.value for cell in sheet[1]] == KOC_EXPORT_COLUMNS
    assert "なみかりちゃんねる😀" in [cell.value for cell in sheet["B"]]
    assert sheet["A2"].number_format == "@"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None
    target_row = next(
        row for row in range(2, sheet.max_row + 1) if sheet.cell(row, 1).value == "920000"
    )
    assert sheet.cell(target_row, 7).hyperlink.target == "https://www.youtube.com/@creator"
    assert sheet.cell(target_row, 8).value == 123456
    assert sheet.cell(target_row, 8).number_format == "#,##0"


def test_contract_types_can_be_replaced_for_monthly_changes(tmp_path):
    repository = KOCRepository(tmp_path / "koc.db")
    record = repository.create(
        user_id="900006",
        koc_name="monthly-contract-creator",
        contract_types=["YTB", "MAY_YTB"],
    )

    updated = repository.update(
        record.id,
        user_id=record.user_id,
        koc_name=record.koc_name,
        creator_category=record.creator_category,
        contract_types=["TT"],
        homepage_url=record.homepage_url,
        follower_count=record.follower_count,
        active=record.active,
        note=record.note,
    )

    assert updated.contract_types == ("TT",)
    assert KOCRepository(tmp_path / "koc.db").get(record.id).contract_types == ("TT",)
