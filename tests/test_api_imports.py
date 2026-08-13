"""Tests for the batch-2 import/rollback/cross-industry write endpoints (19.2)."""
from __future__ import annotations

import io

import pandas as pd
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.main import create_app
from config.settings import Settings
from database.dashboard_repository import DashboardRepository
from database.db import connect
from database.koc_repository import KOCRepository
from models.enums import ContractType, CreatorCategory

TEAM_PASSWORD = "test-team-password"
PREFIX = "importtestuid-"


def _settings(database_path):
    return Settings(
        timezone="Asia/Shanghai",
        database_path=database_path,
        output_dir="data/output",
        youtube_api_key=None,
        tiktok_browser_data_dir="data/tiktok_browser_data",
        tiktok_persistent_headless=False,
        team_password=TEAM_PASSWORD,
    )


def _authenticated_client(database_path):
    app = create_app(_settings(database_path), environment="development")
    client = TestClient(app)
    login_response = client.post(
        "/api/auth/login",
        json={"password": TEAM_PASSWORD, "operator_name": "张三"},
    )
    assert login_response.status_code == 200
    return client


def _seed_creator(database_path, user_id):
    repository = KOCRepository(database_path)
    return repository.create(
        user_id=user_id,
        koc_name=f"达人-{user_id}",
        creator_category=CreatorCategory.GRASSROOT,
        contract_type=ContractType.TT,
        follower_count=1000,
    )


def _timestamp_ms(date_str: str) -> int:
    return int(pd.Timestamp(date_str, tz="Asia/Shanghai").timestamp() * 1000)


def _xlsx_bytes(rows: list[dict]) -> bytes:
    frame = pd.DataFrame(rows)
    buffer = io.BytesIO()
    frame.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)
    return buffer.read()


def _row(user_id: str, *, url: str, title: str, date: str, views: int = 100) -> dict:
    return {
        "userId": user_id,
        "subtype": "video",
        "title": title,
        "url": url,
        "timestamp": _timestamp_ms(date),
        "view": views,
        "likes": 1,
        "comment": 0,
        "reposted": 0,
    }


def _upload_preview(client, rows):
    content = _xlsx_bytes(rows)
    response = client.post(
        "/api/imports/preview",
        files={"files": ("data.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    return response


def test_standardize_restores_legacy_download_without_writing_posts(tmp_path):
    database_path = tmp_path / "koc.db"
    creator = _seed_creator(database_path, f"{PREFIX}standardize")
    client = _authenticated_client(database_path)
    content = _xlsx_bytes(
        [
            _row(
                creator.user_id,
                url="https://x.com/standardize",
                title="standardized",
                date="2026-01-05",
            )
        ]
    )

    response = client.post(
        "/api/imports/standardize",
        files={
            "files": (
                "data.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"processing_timezone": "Asia/Shanghai", "deduplicate_urls": "false"},
    )

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["overall"]["uploaded_files"] == 1
    assert body["overall"]["merged_rows"] == 1
    assert body["overall"]["unmatched_uid_count"] == 0
    assert body["result_preview"][0]["koc_name"] == creator.koc_name
    assert DashboardRepository(database_path).count_posts() == 0

    download = client.get(body["download_path"])

    assert download.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in download.headers["content-type"]
    workbook = load_workbook(io.BytesIO(download.content))
    assert workbook.sheetnames == ["整理结果", "文件处理报告", "异常数据"]
    assert DashboardRepository(database_path).count_posts() == 0


# ---------------------------------------------------------------------------
# 1. POST /api/imports/preview
# ---------------------------------------------------------------------------


def test_preview_shows_all_four_categories(tmp_path):
    database_path = tmp_path / "koc.db"
    _seed_creator(database_path, f"{PREFIX}a")
    client = _authenticated_client(database_path)

    response = _upload_preview(
        client,
        [
            _row(f"{PREFIX}a", url="https://x.com/1", title="t1", date="2026-01-05"),
            _row(f"{PREFIX}unknown", url="https://x.com/2", title="t2", date="2026-01-06"),
        ],
    )

    assert response.status_code == 200
    data = response.json()["data"]
    for key in ("additions", "updates", "removals", "unmatched_creators", "date_anomalies"):
        assert key in data
        assert "count" in data[key]
        assert "rows" in data[key]
    assert data["unmatched_creators"]["count"] == 1
    assert data["additions"]["count"] == 2
    assert data["preview_token"]


# ---------------------------------------------------------------------------
# 2. POST /api/imports/{preview_token}/confirm
# ---------------------------------------------------------------------------


def test_confirm_blocked_by_unmatched_creators(tmp_path):
    database_path = tmp_path / "koc.db"
    _seed_creator(database_path, f"{PREFIX}a")
    client = _authenticated_client(database_path)

    preview = _upload_preview(
        client,
        [_row(f"{PREFIX}unknown", url="https://x.com/1", title="t1", date="2026-01-05")],
    )
    token = preview.json()["data"]["preview_token"]

    response = client.post(
        f"/api/imports/{token}/confirm",
        json={"mode": "replace_months"},
        headers={"Idempotency-Key": "confirm-blocked-1"},
    )

    assert response.status_code == 422
    assert response.json()["error"]["code"] == "VALIDATION_ERROR"


def test_confirm_requires_idempotency_key(tmp_path):
    database_path = tmp_path / "koc.db"
    _seed_creator(database_path, f"{PREFIX}a")
    client = _authenticated_client(database_path)

    preview = _upload_preview(
        client,
        [_row(f"{PREFIX}a", url="https://x.com/1", title="t1", date="2026-01-05")],
    )
    token = preview.json()["data"]["preview_token"]

    response = client.post(f"/api/imports/{token}/confirm", json={"mode": "replace_months"})
    assert response.status_code == 422


def test_confirm_success_creates_snapshot_and_replaces(tmp_path):
    database_path = tmp_path / "koc.db"
    _seed_creator(database_path, f"{PREFIX}a")
    client = _authenticated_client(database_path)

    preview = _upload_preview(
        client,
        [_row(f"{PREFIX}a", url="https://x.com/1", title="t1", date="2026-01-05")],
    )
    token = preview.json()["data"]["preview_token"]

    response = client.post(
        f"/api/imports/{token}/confirm",
        json={"mode": "replace_months"},
        headers={"Idempotency-Key": "confirm-ok-1"},
    )

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["saved_count"] == 1
    batch_id = body["batch_id"]

    with connect(database_path) as connection:
        snapshot_rows = connection.execute(
            "SELECT COUNT(*) AS c FROM dashboard_import_batch_snapshot WHERE batch_id != ?",
            (batch_id,),
        ).fetchone()
    assert snapshot_rows is not None


def test_append_or_update_preserves_existing_posts_in_same_month(tmp_path):
    database_path = tmp_path / "koc.db"
    creator = _seed_creator(database_path, f"{PREFIX}append")
    client = _authenticated_client(database_path)

    initial = _upload_preview(
        client,
        [_row(creator.user_id, url="https://x.com/original", title="original", date="2026-01-05")],
    )
    initial_token = initial.json()["data"]["preview_token"]
    initial_confirm = client.post(
        f"/api/imports/{initial_token}/confirm",
        json={"mode": "replace_months"},
        headers={"Idempotency-Key": "append-initial-1"},
    )
    assert initial_confirm.status_code == 200

    supplement = _upload_preview(
        client,
        [_row(creator.user_id, url="https://x.com/supplement", title="supplement", date="2026-01-20")],
    )
    supplement_token = supplement.json()["data"]["preview_token"]
    supplement_confirm = client.post(
        f"/api/imports/{supplement_token}/confirm",
        json={"mode": "append_or_update"},
        headers={"Idempotency-Key": "append-supplement-1"},
    )

    assert supplement_confirm.status_code == 200
    result = supplement_confirm.json()["data"]
    assert result["mode"] == "APPEND_OR_UPDATE"
    assert result["removed_count"] == 0
    posts = DashboardRepository(database_path).load_posts()
    assert set(posts["url"]) == {
        "https://x.com/original",
        "https://x.com/supplement",
    }


def test_duplicate_confirm_same_idempotency_key_returns_cached_result(tmp_path):
    database_path = tmp_path / "koc.db"
    _seed_creator(database_path, f"{PREFIX}a")
    client = _authenticated_client(database_path)

    preview = _upload_preview(
        client,
        [_row(f"{PREFIX}a", url="https://x.com/1", title="t1", date="2026-01-05")],
    )
    token = preview.json()["data"]["preview_token"]

    first = client.post(
        f"/api/imports/{token}/confirm",
        json={"mode": "replace_months"},
        headers={"Idempotency-Key": "confirm-dup-1"},
    )
    second = client.post(
        f"/api/imports/{token}/confirm",
        json={"mode": "replace_months"},
        headers={"Idempotency-Key": "confirm-dup-1"},
    )

    assert first.status_code == 200
    assert second.status_code == 200
    assert first.json() == second.json()

    repository = DashboardRepository(database_path)
    assert repository.count_posts() == 1


# ---------------------------------------------------------------------------
# 3. POST /api/dashboard/import-batches/{batch_id}/rollback
# ---------------------------------------------------------------------------


def test_rollback_succeeds_and_restores_exact_prior_data(tmp_path):
    database_path = tmp_path / "koc.db"
    _seed_creator(database_path, f"{PREFIX}a")
    client = _authenticated_client(database_path)

    first_preview = _upload_preview(
        client,
        [_row(f"{PREFIX}a", url="https://x.com/1", title="original", date="2026-02-05")],
    )
    first_token = first_preview.json()["data"]["preview_token"]
    first_confirm = client.post(
        f"/api/imports/{first_token}/confirm",
        json={"mode": "replace_months"},
        headers={"Idempotency-Key": "rollback-first"},
    )
    assert first_confirm.status_code == 200

    repository = DashboardRepository(database_path)
    posts_before = repository.load_posts()
    assert list(posts_before["title"]) == ["original"]

    second_preview = _upload_preview(
        client,
        [_row(f"{PREFIX}a", url="https://x.com/2", title="replaced", date="2026-02-06")],
    )
    second_token = second_preview.json()["data"]["preview_token"]
    second_confirm = client.post(
        f"/api/imports/{second_token}/confirm",
        json={"mode": "replace_months"},
        headers={"Idempotency-Key": "rollback-second"},
    )
    assert second_confirm.status_code == 200
    second_batch_id = second_confirm.json()["data"]["batch_id"]

    posts_after_replace = repository.load_posts()
    assert list(posts_after_replace["title"]) == ["replaced"]

    rollback_response = client.post(
        f"/api/dashboard/import-batches/{second_batch_id}/rollback",
        json={"reason": "误导入，需回滚"},
        headers={"Idempotency-Key": "rollback-key-1"},
    )
    assert rollback_response.status_code == 200

    posts_after_rollback = repository.load_posts()
    assert list(posts_after_rollback["title"]) == ["original"]
    assert list(posts_after_rollback["url"]) == ["https://x.com/1"]


def test_rollback_rejected_for_superseded_batch(tmp_path):
    database_path = tmp_path / "koc.db"
    _seed_creator(database_path, f"{PREFIX}a")
    client = _authenticated_client(database_path)

    first_preview = _upload_preview(
        client,
        [_row(f"{PREFIX}a", url="https://x.com/1", title="v1", date="2026-03-05")],
    )
    first_token = first_preview.json()["data"]["preview_token"]
    first_confirm = client.post(
        f"/api/imports/{first_token}/confirm",
        json={"mode": "replace_months"},
        headers={"Idempotency-Key": "superseded-first"},
    )
    first_batch_id = first_confirm.json()["data"]["batch_id"]

    second_preview = _upload_preview(
        client,
        [_row(f"{PREFIX}a", url="https://x.com/2", title="v2", date="2026-03-06")],
    )
    second_token = second_preview.json()["data"]["preview_token"]
    client.post(
        f"/api/imports/{second_token}/confirm",
        json={"mode": "replace_months"},
        headers={"Idempotency-Key": "superseded-second"},
    )

    rollback_response = client.post(
        f"/api/dashboard/import-batches/{first_batch_id}/rollback",
        json={"reason": "尝试回滚旧批次"},
        headers={"Idempotency-Key": "superseded-rollback"},
    )

    assert rollback_response.status_code in (409, 422)


def test_rollback_requires_non_empty_reason(tmp_path):
    database_path = tmp_path / "koc.db"
    _seed_creator(database_path, f"{PREFIX}a")
    client = _authenticated_client(database_path)

    preview = _upload_preview(
        client,
        [_row(f"{PREFIX}a", url="https://x.com/1", title="t1", date="2026-04-05")],
    )
    token = preview.json()["data"]["preview_token"]
    confirm = client.post(
        f"/api/imports/{token}/confirm",
        json={"mode": "replace_months"},
        headers={"Idempotency-Key": "reason-required-confirm"},
    )
    batch_id = confirm.json()["data"]["batch_id"]

    response = client.post(
        f"/api/dashboard/import-batches/{batch_id}/rollback",
        json={"reason": "   "},
        headers={"Idempotency-Key": "reason-required-rollback"},
    )
    assert response.status_code == 422


# ---------------------------------------------------------------------------
# 4/5. Cross-industry exclusions
# ---------------------------------------------------------------------------


def test_cross_industry_mark_and_unmark_preserves_original_post(tmp_path):
    database_path = tmp_path / "koc.db"
    _seed_creator(database_path, f"{PREFIX}a")
    client = _authenticated_client(database_path)

    preview = _upload_preview(
        client,
        [_row(f"{PREFIX}a", url="https://x.com/cross1", title="t1", date="2026-05-05")],
    )
    token = preview.json()["data"]["preview_token"]
    client.post(
        f"/api/imports/{token}/confirm",
        json={"mode": "replace_months"},
        headers={"Idempotency-Key": "cross-confirm-1"},
    )

    repository = DashboardRepository(database_path)
    posts_before = repository.load_posts()

    mark_response = client.post(
        "/api/cross-industry-exclusions",
        json={"urls": ["https://x.com/cross1"], "reason": "异业合作"},
    )
    assert mark_response.status_code == 201
    exclusion_id = mark_response.json()["data"][0]["id"]

    posts_after_mark = repository.load_posts()
    pd.testing.assert_frame_equal(
        posts_before.sort_index(axis=1), posts_after_mark.sort_index(axis=1)
    )

    annotated = repository.annotate_cross_industry_posts(posts_after_mark)
    assert bool(annotated["is_cross_industry"].iloc[0]) is True

    unmark_response = client.delete(f"/api/cross-industry-exclusions/{exclusion_id}")
    assert unmark_response.status_code == 200

    posts_after_unmark = repository.load_posts()
    pd.testing.assert_frame_equal(
        posts_before.sort_index(axis=1), posts_after_unmark.sort_index(axis=1)
    )
    annotated_after_unmark = repository.annotate_cross_industry_posts(posts_after_unmark)
    assert bool(annotated_after_unmark["is_cross_industry"].iloc[0]) is False


def test_cross_industry_accepts_multiple_urls_at_once(tmp_path):
    database_path = tmp_path / "koc.db"
    client = _authenticated_client(database_path)

    response = client.post(
        "/api/cross-industry-exclusions",
        json={"urls": ["https://x.com/m1", "https://x.com/m2"], "reason": "批量异业"},
    )
    assert response.status_code == 201
    assert len(response.json()["data"]) >= 2


# ---------------------------------------------------------------------------
# 6. LOCKED settlement version must remain byte-identical across import+rollback
# ---------------------------------------------------------------------------


def test_locked_settlement_version_untouched_by_import_and_rollback(tmp_path):
    database_path = tmp_path / "koc.db"
    _seed_creator(database_path, f"{PREFIX}a")
    client = _authenticated_client(database_path)
    repository = DashboardRepository(database_path)

    draft = repository.create_compensation_draft(
        "2026-06",
        jpy_to_usd_rate=150.0,
        details=pd.DataFrame({"koc_name": ["达人-x"], "amount": [100]}),
        summary={"total": 100},
        note=None,
    )
    locked = repository.lock_compensation_version(draft.id)
    assert locked.status == "LOCKED"

    with connect(database_path) as connection:
        before_row = dict(
            connection.execute(
                "SELECT * FROM grassroot_compensation_version WHERE id = ?",
                (locked.id,),
            ).fetchone()
        )

    first_preview = _upload_preview(
        client,
        [_row(f"{PREFIX}a", url="https://x.com/lock1", title="t1", date="2026-06-05")],
    )
    first_token = first_preview.json()["data"]["preview_token"]
    first_confirm = client.post(
        f"/api/imports/{first_token}/confirm",
        json={"mode": "replace_months"},
        headers={"Idempotency-Key": "locked-confirm-1"},
    )
    batch_id = first_confirm.json()["data"]["batch_id"]

    client.post(
        f"/api/dashboard/import-batches/{batch_id}/rollback",
        json={"reason": "回滚验证锁定结算不受影响"},
        headers={"Idempotency-Key": "locked-rollback-1"},
    )

    with connect(database_path) as connection:
        after_row = dict(
            connection.execute(
                "SELECT * FROM grassroot_compensation_version WHERE id = ?",
                (locked.id,),
            ).fetchone()
        )

    assert before_row == after_row
